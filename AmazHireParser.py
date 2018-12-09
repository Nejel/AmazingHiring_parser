# coding: utf-8

#TODO: Разобраться с глобалами
#TODO: Зафиксить баги с неподтягиваемым экселем
#TODO: Добавить выбор поисковых сценариев в начале
#TODO: Убрать всякий юникод-булшит из вывода в терминал
#TODO: Добавить поиск по pipl.com
#TODO: Добавить поиск по hh.ru

import re
import sqlite3
import os
import openpyxl
import time
from selenium import webdriver
from selenium.webdriver.common.keys import Keys

def OpenBrowser():
    global browser # открываем броузер
    browser.get('https://search.amazinghiring.com/login/?next=/')
    input('It\'s time to accept cookies banner! Please, go to browser, press \'Accept\'., then go bach here and press Enter')
    # логинимся
    Emailinsert = browser.find_element_by_css_selector('#app > div > div:nth-child(3) > div > div.Login-centerBlock___3BZag > form > div:nth-child(4) > div:nth-child(1) > input')
    Passwordinsert = browser.find_element_by_css_selector('#app > div > div:nth-child(3) > div > div.Login-centerBlock___3BZag > form > div:nth-child(4) > div:nth-child(2) > input')
    Emailinsert.send_keys(username)
    Passwordinsert.send_keys(password)
    time.sleep(0.2)
    Passwordinsert.send_keys(Keys.ENTER)
    time.sleep(3)
    ReadFolderAndFiles()

def ReadFolderAndFiles():
    global cwd
    global cwdstr

    for x in (os.listdir(cwd)):
        filename = cwdstr + '\\' + x # собираем конструктор пути файла
        print(filename)
        if filename.endswith(".xlsx") == True:
            #print(getText(filename))
            print(x)
            content.append(ExcelWorks(x)) # list # добавляем в конец листа
            #fullcontent = ''.join(content)
        else: pass
    return content

def ExcelWorks(x):
    global cur
    global conn
    global linktoprofile
    global skills
    global row2


    try:
        workbook = openpyxl.load_workbook(x)
        print(workbook.sheetnames[0])
        sheet = workbook.sheetnames[0]
        worksheet = workbook.get_sheet_by_name(sheet)
    except:
        input("It looks like XLSX is open, please close it and press Enter", )
        workbook = openpyxl.load_workbook(x)
        print(workbook.sheetnames[0])
        worksheet = workbook.get_sheet_by_name(sheet)
        #worksheet = workbook.sheetnames[0]
    print('Looking at the list', sheet)
    row_count = worksheet.max_row
    print('Count of the emails', (int(row_count)-1))
    imax = int(row_count)
    for i in range (2, imax):
        candicemail = worksheet.cell(row = i, column = 1)
        #print(candicemail.value)
        print("Progress " + str(i) + "/" + str(imax) + " with email " + candicemail.value)
        candicemail = str(candicemail.value)
        cur.execute('SELECT count FROM Counts WHERE email = ? ', (candicemail,))
        row = cur.fetchone()
        #linktoprofile = []
        if row is None:
            Search(candicemail)
            #print('before insertions', linktoprofile)
            # insert linktoprofile to xlsx
            #print(type(linktoprofile))
            worksheet.cell(row = i, column = 2).value = linktoprofile
            worksheet.cell(row = i, column = 3).value = skills
            workbook.save(x)
        else:
            #print('It looks like we have alredy found', candicemail)
            cur.execute('UPDATE Counts SET count = count + 1 WHERE email = ? ', (candicemail,))
            conn.commit()
            #print('before insertions ELSE part', linktoprofile)
            GetFromDBtoXLSX(candicemail, i)
            worksheet.cell(row = i, column = 2).value = str(row2[0])
            worksheet.cell(row = i, column = 3).value = str(row2[1])
            workbook.save(x)

def Search(candicemail):
    global browser
    global linktoprofile
    global skills
    candicemailconstr = str(re.sub(r'@', '%40', candicemail))
    URLconstr = ('https://search.amazinghiring.com/profiles/?q=booleanText[0]:' + candicemailconstr)
    browser.get(URLconstr)
    time.sleep(4)
    try:
        linktoprofile = browser.find_element_by_xpath("//a[contains(@href,'profile')]").get_attribute("href")
        print('I found some guy', linktoprofile)
        try:
            skills = browser.find_element_by_xpath("//*[contains(@class,'Skills-skill__skill')]").text
            print('skills found: ', skills)
        except:
            print('unluck to find any skills')#
            skills = 'unluck to find any skills'
        PutToDB(linktoprofile, candicemail, skills)
    except:
        print('Nothing was found')
        linktoprofile = 'Nothing was found'
        skills = 'unluck to find any skills'
    return linktoprofile #, skills

def PutToDB(linktoprofile, candicemail, skills):
    global cur
    global conn
    print('Trying to put him in DB')
    cur.execute('SELECT amazeprofile FROM Counts WHERE email = ? ', (candicemail,))
    row = cur.fetchone()
    try:
        #print('Im trying', linktoprofile)
        cur.execute('INSERT INTO Counts (email, count, amazeprofile, skill) VALUES (?, 1, ?, ?)', (candicemail, linktoprofile, skills)) # что-то здесь нечисто
        #print('After this line')
        conn.commit()
        #print('After conn commit line')
    except:
        print('Something went wrong in PutToDB')
    return linktoprofile


def GetFromDBtoXLSX(candicemail, i):
    global cur
    global conn
    global row2
    print('Trying to take him from DB')
    cur.execute('SELECT amazeprofile, skill FROM Counts WHERE email = ? ', (candicemail,))
    row2 = cur.fetchone()
    try:
        #print('Im trying to GetFromDBtoXLSX %s') % (row2)
        #print("Hi blea" + str(row2))
        #print('before insertions EPTABLEA')
        # insert linktoprofile to xlsx
        # workbook = openpyxl.load_workbook('Testlist.xlsx')
        # sheet = workbook["Sheet1"]
        #print(row2[0])
        #print(row2[1])
        #print(type(row2[0]))
        #print(i)
        # sheet.cell(row = i, column = 2).value = "TESTSTRING" #str(row[0])
        # sheet.cell(row = i, column = 3).value = "TESTSTRING2" #str(row[1])
        # workbook.save('Testlist.xlsx')
        #print('Something was good?')
        pass
    except:
        pass
        #print('Something went wrong in GetFromDBtoXLSX')

print("Mr Neko is ready to help ya (=^‥^)/’`")
username = input('Please enter Am-Hiring USERNAME:', )
password = input('Am-Hiring PASSWD:', )
print("Be patient, wait a second =^._.^= ∫")
linktoprofile = 'Nothing was found'
skills = 'skills wasnt found'
cwd = os.chdir(os.getcwd() + '\\workdir')
cwdstr = str(cwd)
row2 = []

# This part is not using anymore
# whattofind = open('whattofind.txt', 'r')
# whattofindfile = whattofind.read()
# print('I\'am looking for:', whattofindfile.split(','))
# whattofindfilesearch = whattofindfile.split(',')

content = []
x = []
conn = sqlite3.connect('db.sqlite')
cur = conn.cursor()
#cur.execute('DROP TABLE IF EXISTS Counts')
cur.execute('''CREATE TABLE IF NOT EXISTS Counts (email TEXT, count INTEGER, amazeprofile TEXT, skill TEXT)''')
browser = webdriver.Chrome()
OpenBrowser()
